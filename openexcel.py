import openpyxl

# Open the workbook via load_workbook
wb = openpyxl.load_workbook('example.xlsx')
wb1 = openpyxl.load_workbook('example_one.xlsx')

print(type(wb))

print(wb.sheetnames) # it will print all the sheetnames of workbook

sheet = wb['Sheet1'] # to access a worksheet
sheet_one = wb1['Sheet1']

print(sheet['A1'].value) # get value of a cell

c = sheet['B1']
print(c.value)

sheet_one.cell(row = 10, column = 10).value = sheet.cell(row=1, column=1).value # update a cell value
wb.save('example.xlsx')
wb1.save('example_one.xlsx')

print(sheet.title)

activeSheet = wb.active  # to access the active sheet

